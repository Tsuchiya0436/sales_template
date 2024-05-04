from openpyxl import Workbook
import calendar
import datetime

date = datetime.datetime.now()
now_year = date.year
now_month = date.month-1
wareki = now_year % 100 - 18
n = calendar.monthrange(now_year, now_month)[1]
wdays = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']
category = ['年月日', '曜日', '組数', '客数', '売上']

wb = Workbook()
ws = wb.active
ws.title = 'Sheet1'
sheet = ws

for i in range(len(category)):
    sheet.cell(row=1, column=i+1, value=category[i])

row = 2
col = 1

for i in range(1, n+1):
    wd = datetime.date(now_year, now_month, i).weekday()
    wday = wdays[wd]
    cell = sheet.cell(row=row, column=col, value=datetime.date(now_year, now_month, i))
    cell = sheet.cell(row=row, column=col+1, value=wday)
    row += 1

wb.save(f'R{wareki}-{now_month}.xlsx')
